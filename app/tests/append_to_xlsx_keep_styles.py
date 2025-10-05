import pandas as pd
from openpyxl import load_workbook
import os

os.chdir("C:/Users/hoend008/OneDrive - WageningenUR/DEVELOPMENT/mc-api/app/tests")


# Example DataFrame
df = pd.DataFrame({
    "firstname": ["Alice", "Bob", "Charlie"],
    "lastname": ["Smith", "Jones", "Brown"],
    "age": [25, 30, 35]
})

# Path to the existing Excel file
file_path = "people.xlsx"

# Load the existing workbook
book = load_workbook(file_path)

# Select the sheet you want to append to
sheet = book["Sheet1"]

# --- Find last row that actually has data ---
last_data_row = 0
for row in sheet.iter_rows(values_only=True):
    if any(cell is not None and cell != "" for cell in row):
        last_data_row += 1

next_row = last_data_row + 1

# --- Append the dataframe starting right after the last data row ---
for r in df.itertuples(index=False, name=None):
    for col_index, value in enumerate(r, start=1):
        sheet.cell(row=next_row, column=col_index, value=value)
    next_row += 1

# Save the workbook
book.save(file_path)


