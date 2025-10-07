from fastapi import status, APIRouter, Depends, HTTPException
import pandas as pd
import numpy as np
from typing import List
from schemas.schemas import MCdata, MCdataIn, ReturnMessage, MCdataOut, Person
from DB.PostgresDatabasev2 import PostgresDatabase
from DB.DBcredentials import DB_USER, DB_PASSWORD, DB_NAME
from utils.oauth2 import get_current_user
from utils.column_conversion import cols_conversion_dict
from utils.datatype_conversion import datatype_conversion_dict

from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile
import os

import warnings
warnings.filterwarnings('ignore')

router = APIRouter(
    prefix="/mcdata",
    tags=["mcdata"])

@router.get('/', response_model=List[MCdataOut])
def get_user(current_user: int = Depends(get_current_user)):
    with PostgresDatabase(DB_NAME, DB_USER, DB_PASSWORD, realdictcursor=True) as db:
        db.execute("""SELECT * FROM mc.tabel_test""")
        mcdata = db.fetchall()
    if not mcdata:
        raise HTTPException(status.HTTP_404_NOT_FOUND,
                            detail="MC data not found")
    return mcdata  

""" Post MC data to store in DB """
@router.post("/", status_code=status.HTTP_201_CREATED, response_model=ReturnMessage)
def insert_mcdata(mcdata: List[MCdataIn], current_user: int = Depends(get_current_user)) -> ReturnMessage:
  print("inserting data...")
  """
  CREATE DATAFRAME FROM mcdata
  """
  df = pd.DataFrame([model.dict() for model in mcdata])

  """
  REMOVE ROWS WITHOUT team_id
  """
  df = df.query("team_id == team_id")

  """
  RENAME COLUMNS IN df
  """
  df = df.rename(columns=cols_conversion_dict)

  """
  CONVERT DATATYPE
  """
  for col in df.columns:
      # get datatype
      dt = datatype_conversion_dict[col]
      
      # set datatype
      if dt == 'date':
          #df[col] = df[col].dt.date
          df[col] = pd.to_datetime(df[col])
      else:
          # set datatype
          df[col] = df[col].astype(dt)

  """
  REPLACE NAN WITH EMPTY STRING
  """
  df = df.replace('nan', '')

  """
  ADD MISSING COLUMNS
  """
  #for col in datatype_conversion_dict.keys():
  #    if col not in df.columns:
  #        df[col] = np.nan

  """
  TEMPORARY
  """
  # df['identifier'] = 1

  """
  SAVE TO DB
  """
  print("to DB...")
  with PostgresDatabase(DB_NAME, DB_USER, DB_PASSWORD) as db:
      db.insert_update(df.copy(), 
                       'mc.tabel_test', 
                       update_existing=['identifier', 'team_id'],
                       text_output=False)   

      return {"msg": "Data successfully updated"}
  




""" Helper function """
# --- Helper function to find last row with actual data ---
def get_last_data_row(sheet):
    last_data_row = 0
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None and cell != "" for cell in row):
            last_data_row += 1
    return last_data_row

""" Appends data to Excel and then returns the Excel for download """
@router.post("/download")
def append_people(mcdata: List[MCdataIn]):

    cwd = os.getcwd()
    path = os.path.join(cwd, "tests/output_MC_database.xlsx")
    file_path = os.path.abspath(path)

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Excel file not found")

    # get data from DB
    with PostgresDatabase(DB_NAME, DB_USER, DB_PASSWORD) as db:
        df = db.querydf("SELECT * FROM mc.vw_tabel_including_archive ORDER BY team_id, identifier, created_at DESC;")
    # Convert input to DataFrame for convenience
    # df = pd.DataFrame([p.dict() for p in mcdata])

    # reorder columns
    # df = df[datatype_conversion_dict.keys()]

    """
    CONVERT DATATYPE
    """
    for col in df.columns:
        # get datatype
        dt = datatype_conversion_dict[col]
        
        # set datatype
        if dt == 'date':
            #df[col] = df[col].dt.date
            df[col] = pd.to_datetime(df[col])
            if col not in ['created_at', 'updated_at']:
                df[col] = df[col].dt.strftime('%d-%m-%Y').str.replace(r'\b0', '', regex=True)
        else:
            # set datatype
            df[col] = df[col].astype(dt)

    # Load the workbook and sheet
    book = load_workbook(file_path)
    sheet = book["Output"]

    # Find the last data row (ignores formatting-only rows)
    next_row = get_last_data_row(sheet) + 1

    # Append data row by row
    for r in df.itertuples(index=False, name=None):
        for col_index, value in enumerate(r, start=1):
            if not pd.isna(value) and value is not None and value not in ['nan', 'None']:
                sheet.cell(row=next_row, column=col_index, value=value)
        next_row += 1

    # row styling (grey color for each odd numbered row)
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Apply style to every odd-numbered row (excluding header if desired)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # skip header row (start from 2)
        if row[0].row % 2 == 1:  # odd-numbered row
            for cell in row:
                cell.fill = grey_fill

    # Save to a temporary file so we don't overwrite the original
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        temp_path = tmp.name
        book.save(temp_path)

    # Return the file as a download
    return FileResponse(
        temp_path,
        filename="updated_testMC.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
